"""XLSX extractor: one canonical page per sheet, each holding a table block.

Preserves the security posture of the legacy parser: XLSX ingestion requires an
explicit, authoritative ACL scope. Access control is never inferred from a
workbook, so extraction refuses to proceed without non-empty ACL groups even
though the manifest itself never stores them.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from eda.canonical import (
    build_manifest,
    build_table,
    lightweight_provenance,
    make_block,
    native_page_result,
    table_reading_text,
)
from eda.identifiers import page_id as build_page_id, sha256_file
from eda.ingestion_schema import BlockType, ExtractionRoute, TextLayer
from eda.ocr_quality import normalize_ocr_text
from eda.quality import level_for_score, score_text, to_processing_status
from eda.extractors.base import (
    DocumentExtractor,
    ExtractedDocument,
    ExtractionFailed,
    ExtractionOptions,
    ExtractorRoute,
)

PARSER_NAME = "xlsx"
PARSER_VERSION = "1.0"


def _require_acl(acl_groups: tuple[str, ...]) -> None:
    if not acl_groups or not any(isinstance(g, str) and g.strip() for g in acl_groups):
        raise ValueError(
            "XLSX ingestion requires at least one authoritative ACL group; "
            "missing ACL is never treated as public access."
        )


class XlsxExtractor(DocumentExtractor):
    route = ExtractorRoute.XLSX
    supported_extensions = frozenset({".xlsx"})

    def extract(self, path: str | Path, *, options: ExtractionOptions) -> ExtractedDocument:
        _require_acl(options.acl_groups)
        source = Path(path)
        if not source.is_file():
            raise ExtractionFailed("XLSX source is unavailable or unreadable.")
        digest = sha256_file(source)
        try:
            workbook = openpyxl.load_workbook(source, data_only=True, read_only=True)
        except Exception as error:  # noqa: BLE001 - message intentionally path-free
            raise ExtractionFailed("XLSX source could not be opened.") from error

        sheet_names = [name for name in workbook.sheetnames if not name.startswith("_")]
        manifest = build_manifest(
            path=source,
            tenant_id=options.tenant_id,
            logical_document_key=options.logical_document_key,
            source_name=options.source_name,
            source_type="xlsx",
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            parser_name=PARSER_NAME,
            parser_version=PARSER_VERSION,
            total_pages=max(len(sheet_names), 1),
            digest=digest,
        )
        provenance = lightweight_provenance(
            digest=digest, pipeline_name=PARSER_NAME, pipeline_version=PARSER_VERSION,
        )

        pages = []
        page_index = 0
        for name in sheet_names:
            worksheet = workbook[name]
            rows = [
                [("" if value is None else str(value)) for value in row]
                for row in worksheet.iter_rows(values_only=True)
            ]
            rows = [row for row in rows if any(cell.strip() for cell in row)]
            table = build_table(rows, caption=name) if rows else None
            page_id = build_page_id(str(manifest.revision_id), page_index)
            blocks = []
            if table is not None:
                blocks.append(
                    make_block(
                        page_id=page_id, page_number=page_index + 1, block_type=BlockType.TABLE,
                        text=table_reading_text(table), reading_order=0,
                        source_layer=TextLayer.NATIVE, extraction_route=ExtractionRoute.NATIVE,
                        table=table, metadata={"sheet": name},
                    )
                )
            native_text = normalize_ocr_text("\n".join(block.text for block in blocks))
            score = score_text(native_text)
            level = level_for_score(score)
            pages.append(
                native_page_result(
                    manifest=manifest, page_index=page_index, native_text=native_text,
                    provenance=provenance, blocks=blocks,
                    processing_status=to_processing_status(level),
                    quality_gate_passed=level.value in {"accepted", "accepted_with_warning"},
                    quality_score=score,
                )
            )
            page_index += 1
        workbook.close()
        return ExtractedDocument.build(manifest=manifest, pages=pages, extractor_route=self.route)
