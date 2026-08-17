import pymupdf
from eda.chunk import split_fixed
from eda.schema import Chunk
import openpyxl
from eda.normalize import persian_ratio
import os
import docx
import pytesseract
from PIL import Image

from eda.adaptive_ocr import AdaptiveOCRConfig, AdaptivePDFExtractor
from eda.ingestion_adapter import extraction_to_chunks

# Tesseract is a system binary, not a Python package, so its location is a
# property of the machine rather than of this repo. Default to whatever
# `tesseract` is on PATH — correct inside the container, where the image installs
# it — and let TESSERACT_CMD override on hosts that keep it somewhere else
# (a Windows dev box, typically). Nothing machine-specific is hardcoded here.
_TESSERACT_CMD = os.getenv("TESSERACT_CMD")
if _TESSERACT_CMD:
    pytesseract.pytesseract.tesseract_cmd = _TESSERACT_CMD



PDF_EXTRACTION_MODES = frozenset({"native", "adaptive"})


def resolve_pdf_extraction_mode(explicit_mode: str | None = None) -> str:
    """Read only the explicit argument or EDA_PDF_EXTRACTION_MODE."""
    mode = explicit_mode if explicit_mode is not None else os.getenv("EDA_PDF_EXTRACTION_MODE", "native")
    if not isinstance(mode, str):
        raise ValueError("EDA_PDF_EXTRACTION_MODE must be 'native' or 'adaptive'.")
    mode = mode.strip().casefold()
    if mode not in PDF_EXTRACTION_MODES:
        raise ValueError("EDA_PDF_EXTRACTION_MODE must be 'native' or 'adaptive'.")
    return mode


def _parse_pdf_native(path, tenant_id, source, acl_groups=None, lang="en",
                      page_start=0, page_end=None):
    if acl_groups is None:
        acl_groups = []

    doc = pymupdf.open(path)
    if page_end is None:
        page_end = len(doc)

    chunks = []
    for page_num, page in enumerate(doc):
        if page_num < page_start:
            continue
        if page_num >= page_end:
            break
        text = page.get_text()
        for i, piece in enumerate(split_fixed(text)):
            chunks.append(Chunk(
                text=piece,
                lang=lang,
                tenant_id=tenant_id,
                source=source,
                source_type="pdf",
                location={"type": "page", "num": page_num},
                acl_groups=acl_groups,
                chunk_id=f"{source}-{page_num}_{i}",
            ))

    doc.close()
    return chunks


def parse_pdf(path, tenant_id, source, acl_groups=None, lang="en",
              page_start=0, page_end=None, extraction_mode=None,
              logical_document_key=None, adaptive_config=None,
              allow_needs_review=False):
    """Parse a PDF; native remains the default and preserves legacy behavior."""
    mode = resolve_pdf_extraction_mode(extraction_mode)
    # print("PDF extraction mode:", os.getenv("EDA_PDF_EXTRACTION_MODE"), flush=True)
    if mode == "native":
        return _parse_pdf_native(
            path,
            tenant_id,
            source,
            acl_groups=acl_groups,
            lang=lang,
            page_start=page_start,
            page_end=page_end,
        )
    if not logical_document_key:
        raise ValueError("adaptive PDF extraction requires logical_document_key.")
    config = adaptive_config or AdaptiveOCRConfig()
    if not isinstance(config, AdaptiveOCRConfig):
        raise TypeError("adaptive_config must be an AdaptiveOCRConfig instance.")
    extraction = AdaptivePDFExtractor(config).extract(
        path,
        tenant_id=tenant_id,
        logical_document_key=logical_document_key,
        source_name=os.path.basename(source),
        page_start=page_start,
        page_end=page_end,
    )
    return extraction_to_chunks(
        extraction.pages,
        extraction.manifest,
        acl_groups=acl_groups,
        allow_needs_review=allow_needs_review,
    )

def _require_acl_groups(acl_groups) -> list[str]:
    if not acl_groups:
        raise ValueError(
            "XLSX ingestion requires at least one authoritative ACL group; "
            "missing ACL is never treated as public access."
        )

    normalized = []
    for group in acl_groups:
        if not isinstance(group, str) or not group.strip():
            raise ValueError("ACL groups must be non-empty strings.")
        group = group.strip()
        if group not in normalized:
            normalized.append(group)
    return normalized


def parse_xlsx(path, tenant_id, source, acl_groups=None):
    acl_groups = _require_acl_groups(acl_groups)
    wb = openpyxl.load_workbook(path, data_only=True)
    chunks = []

    for sheet_name in wb.sheetnames:
        if sheet_name.startswith("_"):
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        # هدر: اولین ردیف با ۷۰٪ پر
        header_row = None
        for r, row in enumerate(rows):
            filled = sum(1 for v in row if v is not None and str(v).strip())
            if filled >= len(row) * 0.7:
                header_row = r
                break
        if header_row is None:
            continue
        headers = rows[header_row]
        data_rows = rows[header_row + 1:]

        text_col_idx = []
        for col in range(len(headers)):
            lengths = [len(str(row[col])) for row in data_rows
                       if col < len(row) and row[col] is not None]
            if lengths and sum(lengths) / len(lengths) > 40:  # میانگین > ۴۰ کاراکتر
                text_col_idx.append(col)


        # TODO: برای هر ردیف داده، متن رو از ستون‌های متنی بساز
        for r, row in enumerate(data_rows):
            filled = sum(1 for v in row if v is not None and str(v).strip())
            if filled < 2:  # جداکننده رو رد کن
                continue
            parts = []
            for col in text_col_idx:
                if col < len(row) and row[col] is not None and str(row[col]).strip():
                    parts.append(f"{headers[col]}: {row[col]}")
            text = "\n".join(parts)
            if not text:
                continue
            chunks.append(Chunk(
                text=text,
                lang="fa" if persian_ratio(text) > 0.5 else "en",
                tenant_id=tenant_id,
                source=source,
                source_type="xlsx",
                location={
                    "type": "cell",
                    "sheet": sheet_name,
                    "row": header_row + 1 + r + 1,
                    "ref": str(row[0]) if row[0] else None,
                },
                acl_groups=acl_groups,  # ← از پارامتر، نه فایل
                chunk_id=f"{source}-{sheet_name}-{r}",
            ))

        # TODO: Chunk بساز — acl_groups از پارامتر، نه از فایل

    return chunks


def parse_docx(path, tenant_id, source, acl_groups=None, lang="en"):
    if acl_groups is None:
        acl_groups = []
    doc = docx.Document(path)
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    chunks = []
    for i, piece in enumerate(split_fixed(full_text)):
        chunks.append(Chunk(
            text=piece,
            lang="fa" if persian_ratio(piece) > 0.5 else "en",   # ← خودکار
            tenant_id=tenant_id,
            source=source,
            source_type="docx",
            location={"type": "paragraph", "index": i},
            acl_groups=acl_groups,
            chunk_id=f"{source}-{i}",
        ))
    return chunks


def parse_image(path, tenant_id, source, acl_groups=None):
    if acl_groups is None:
        acl_groups = []

    # OCR — هر دو زبان، بذار خودش تشخیص بده
    img = Image.open(path)
    text = pytesseract.image_to_string(img, lang="eng+fas")
    text = text.strip()

    if not text:
        return []      # عکس بدون متن (یا OCR شکست خورد)

    chunks = []
    for i, piece in enumerate(split_fixed(text)):
        chunks.append(Chunk(
            text=piece,
            lang="fa" if persian_ratio(piece) > 0.5 else "en",
            tenant_id=tenant_id,
            source=source,
            source_type="image",
            location={"type": "image", "num": i},
            acl_groups=acl_groups,
            chunk_id=f"{source}-{i}",
        ))
    return chunks


PARSERS = {
    ".pdf": parse_pdf,
    ".xlsx": parse_xlsx,
    ".docx": parse_docx,
    ".png": parse_image,      # ← سه خط برای سه فرمت تصویر
    ".jpg": parse_image,
    ".jpeg": parse_image,
}


def parse_any(path, tenant_id, source, acl_groups=None, **kwargs):
    ext = os.path.splitext(path)[1].lower()
    parser = PARSERS.get(ext)
    if parser is None:
        raise ValueError(f"فرمت پشتیبانی نمی‌شود: {ext}")
    return parser(path, tenant_id, source, acl_groups=acl_groups, **kwargs)
